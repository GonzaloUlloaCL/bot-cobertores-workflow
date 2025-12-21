"""
Attachment Processor - Procesamiento de archivos adjuntos
Maneja Excel (XLS/XLSX), PDF, y CSV
"""

import os
import logging
from typing import List, Dict, Optional
import pandas as pd
from openpyxl import load_workbook
import PyPDF2

# Configurar logging
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)


class AttachmentProcessor:
    """Procesador de archivos adjuntos (Excel, PDF, CSV)"""
    
    def __init__(self):
        """Inicializa el procesador"""
        logger.info("✅ AttachmentProcessor inicializado")
    
    def process_file(self, file_path: str) -> Optional[List[Dict]]:
        """
        Procesa un archivo según su extensión
        
        Args:
            file_path: Ruta al archivo
        
        Returns:
            Lista de diccionarios con datos extraídos o None si falla
        """
        if not os.path.exists(file_path):
            logger.error(f"❌ Archivo no encontrado: {file_path}")
            return None
        
        file_ext = os.path.splitext(file_path)[1].lower()
        
        try:
            if file_ext in ['.xlsx', '.xls']:
                return self.process_excel(file_path)
            elif file_ext == '.pdf':
                return self.process_pdf(file_path)
            elif file_ext == '.csv':
                return self.process_csv(file_path)
            else:
                logger.warning(f"⚠️ Tipo de archivo no soportado: {file_ext}")
                return None
        except Exception as e:
            logger.error(f"❌ Error procesando archivo {file_path}: {e}")
            return None
    
    def process_excel(self, file_path: str) -> List[Dict]:
        """
        Procesa archivo Excel (XLS/XLSX)
        
        Args:
            file_path: Ruta al archivo Excel
        
        Returns:
            Lista de diccionarios con datos extraídos
        """
        logger.info(f"📊 Procesando Excel: {os.path.basename(file_path)}")
        
        results = []
        
        try:
            # Intentar leer con pandas primero (más robusto)
            excel_file = pd.ExcelFile(file_path, engine='openpyxl')
            
            # Procesar todas las hojas
            for sheet_name in excel_file.sheet_names:
                logger.info(f"   📄 Procesando hoja: {sheet_name}")
                
                df = pd.read_excel(file_path, sheet_name=sheet_name, engine='openpyxl')
                
                # Limpiar nombres de columnas
                df.columns = df.columns.str.strip().str.lower()
                
                # Extraer datos
                sheet_data = self._extract_from_dataframe(df, sheet_name)
                results.extend(sheet_data)
            
            logger.info(f"✅ Excel procesado: {len(results)} registros extraídos")
            return results
            
        except Exception as e:
            logger.error(f"❌ Error procesando Excel: {e}")
            # Intentar método alternativo con openpyxl directamente
            return self._process_excel_openpyxl(file_path)
    
    def _process_excel_openpyxl(self, file_path: str) -> List[Dict]:
        """Método alternativo usando openpyxl directamente"""
        try:
            wb = load_workbook(file_path, data_only=True)
            results = []
            
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                
                # Convertir a lista de listas
                data = []
                for row in ws.iter_rows(values_only=True):
                    data.append(row)
                
                # Crear DataFrame
                if len(data) > 1:
                    df = pd.DataFrame(data[1:], columns=data[0])
                    df.columns = df.columns.str.strip().str.lower()
                    sheet_data = self._extract_from_dataframe(df, sheet_name)
                    results.extend(sheet_data)
            
            return results
        except Exception as e:
            logger.error(f"❌ Error con openpyxl: {e}")
            return []
    
    def _extract_from_dataframe(self, df: pd.DataFrame, source: str = "") -> List[Dict]:
        """
        Extrae datos estructurados de un DataFrame
        
        Args:
            df: DataFrame de pandas
            source: Nombre de la fuente (sheet, archivo)
        
        Returns:
            Lista de diccionarios con datos estructurados
        """
        results = []
        
        # Mapeo de nombres de columnas (case-insensitive, flexibles)
        column_mapping = {
            'codigo': ['codigo', 'código', 'codigo_cobertor', 'cod', 'code'],
            'cuartel': ['cuartel', 'quartel', 'sector', 'campo'],
            'hileras': ['hileras', 'hilera', 'rows', 'filas'],
            'largo': ['largo', 'largo_metros', 'largo (m)', 'metros', 'length'],
            'prioridad': ['prioridad', 'priority', 'urgencia', 'nivel']
        }
        
        # Detectar columnas disponibles
        detected_cols = {}
        for field, variations in column_mapping.items():
            for col in df.columns:
                if any(var in str(col).lower() for var in variations):
                    detected_cols[field] = col
                    break
        
        logger.info(f"   🔍 Columnas detectadas: {list(detected_cols.keys())}")
        
        # Iterar sobre filas
        for idx, row in df.iterrows():
            try:
                # Extraer datos con manejo de valores nulos
                codigo = self._get_value(row, detected_cols.get('codigo'))
                cuartel = self._get_value(row, detected_cols.get('cuartel'))
                hileras = self._get_value(row, detected_cols.get('hileras'), cast_type='int')
                largo = self._get_value(row, detected_cols.get('largo'), cast_type='float')
                prioridad = self._get_value(row, detected_cols.get('prioridad'))
                
                # Validar que al menos tengamos código o cuartel
                if not codigo and not cuartel:
                    continue
                
                # Normalizar prioridad
                prioridad_norm = self._normalize_priority(prioridad)
                
                record = {
                    'codigo_cobertor': codigo,
                    'cuartel': str(cuartel) if cuartel else None,
                    'hileras': hileras,
                    'largo_metros': largo,
                    'prioridad': prioridad_norm,
                    'descripcion': f"Registro de {source}" if source else None,
                    'notas': f"Fila {idx + 2}",  # +2 porque Excel empieza en 1 y hay header
                    'urgente': prioridad_norm == 'alta',
                    'origen': 'excel_adjunto'
                }
                
                results.append(record)
                
            except Exception as e:
                logger.warning(f"⚠️ Error procesando fila {idx}: {e}")
                continue
        
        return results
    
    def _get_value(self, row, col_name, cast_type=None):
        """Obtiene valor de una fila manejando nulos y tipos"""
        if col_name is None or col_name not in row.index:
            return None
        
        value = row[col_name]
        
        # Manejar valores nulos de pandas
        if pd.isna(value):
            return None
        
        # Cast a tipo específico
        try:
            if cast_type == 'int':
                return int(float(value))
            elif cast_type == 'float':
                return float(value)
            else:
                return str(value).strip() if value else None
        except (ValueError, TypeError):
            return None
    
    def _normalize_priority(self, value) -> str:
        """Normaliza valores de prioridad"""
        if not value:
            return 'normal'
        
        value_lower = str(value).lower().strip()
        
        if any(word in value_lower for word in ['alta', 'high', 'urgent', 'critica', 'critical']):
            return 'alta'
        elif any(word in value_lower for word in ['baja', 'low']):
            return 'baja'
        else:
            return 'normal'
    
    def process_pdf(self, file_path: str) -> List[Dict]:
        """
        Procesa archivo PDF extrayendo texto
        
        Args:
            file_path: Ruta al PDF
        
        Returns:
            Lista con texto extraído (para posterior procesamiento con IA)
        """
        logger.info(f"📄 Procesando PDF: {os.path.basename(file_path)}")
        
        try:
            with open(file_path, 'rb') as file:
                pdf_reader = PyPDF2.PdfReader(file)
                
                text_content = []
                for page_num in range(len(pdf_reader.pages)):
                    page = pdf_reader.pages[page_num]
                    text = page.extract_text()
                    text_content.append(text)
                
                full_text = '\n'.join(text_content)
                
                logger.info(f"✅ PDF procesado: {len(full_text)} caracteres extraídos")
                
                # Retornar como estructura para posterior procesamiento con IA
                return [{
                    'texto_completo': full_text,
                    'origen': 'pdf_adjunto',
                    'requiere_procesamiento_ia': True
                }]
                
        except Exception as e:
            logger.error(f"❌ Error procesando PDF: {e}")
            return []
    
    def process_csv(self, file_path: str) -> List[Dict]:
        """
        Procesa archivo CSV
        
        Args:
            file_path: Ruta al CSV
        
        Returns:
            Lista de diccionarios con datos extraídos
        """
        logger.info(f"📋 Procesando CSV: {os.path.basename(file_path)}")
        
        try:
            # Intentar con diferentes encodings
            encodings = ['utf-8', 'latin-1', 'iso-8859-1', 'cp1252']
            
            df = None
            for encoding in encodings:
                try:
                    df = pd.read_csv(file_path, encoding=encoding)
                    logger.info(f"   ✅ CSV leído con encoding: {encoding}")
                    break
                except:
                    continue
            
            if df is None:
                logger.error("❌ No se pudo leer el CSV con ningún encoding")
                return []
            
            # Limpiar columnas
            df.columns = df.columns.str.strip().str.lower()
            
            # Extraer datos
            results = self._extract_from_dataframe(df, os.path.basename(file_path))
            
            logger.info(f"✅ CSV procesado: {len(results)} registros extraídos")
            return results
            
        except Exception as e:
            logger.error(f"❌ Error procesando CSV: {e}")
            return []


# Función helper para uso rápido
def process_attachment(file_path: str) -> Optional[List[Dict]]:
    """
    Función helper para procesar un adjunto rápidamente
    
    Usage:
        from data_processing.attachment_processor import process_attachment
        data = process_attachment('path/to/file.xlsx')
    """
    processor = AttachmentProcessor()
    return processor.process_file(file_path)


if __name__ == "__main__":
    # Test del procesador
    print("🧪 Testing Attachment Processor...")
    
    # Crear archivo Excel de prueba
    test_data = {
        'Código': ['COB-003', 'COB-004', 'COB-005'],
        'Cuartel': ['18', '20', '22'],
        'Hileras': [10, 12, 8],
        'Largo (m)': [145.5, 160.0, 130.5],
        'Prioridad': ['NORMAL', 'URGENTE', 'NORMAL']
    }
    
    test_df = pd.DataFrame(test_data)
    test_file = 'test_cobertores.xlsx'
    test_df.to_excel(test_file, index=False, engine='openpyxl')
    
    print(f"📊 Archivo de prueba creado: {test_file}")
    
    # Procesar
    processor = AttachmentProcessor()
    results = processor.process_excel(test_file)
    
    if results:
        print(f"\n✅ TEST EXITOSO! {len(results)} registros extraídos\n")
        for i, record in enumerate(results, 1):
            print(f"Registro {i}:")
            print(f"  - Código: {record['codigo_cobertor']}")
            print(f"  - Cuartel: {record['cuartel']}")
            print(f"  - Hileras: {record['hileras']}")
            print(f"  - Largo: {record['largo_metros']}m")
            print(f"  - Prioridad: {record['prioridad']}")
            print(f"  - Urgente: {record['urgente']}")
            print()
    else:
        print("\n❌ TEST FALLIDO")
    
    # Limpiar archivo de prueba
    import os
    if os.path.exists(test_file):
        os.remove(test_file)
        print(f"🗑️ Archivo de prueba eliminado")